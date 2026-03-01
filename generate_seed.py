#!/usr/bin/env python3
"""
Serviaux - Seed Data Generator
Reads Excel files and generates seed_data.sql for the Android app to preload on first run.

Usage:
    python generate_seed.py

Input files (in data/ directory):
    - Catalogos.xlsx        (Users, Vehicle Types, Colors, Accessories, Part Brands, Services, Motivos, Diagnosticos)
    - Marcas y Modelos.xlsx (Vehicle Brands & Models)
    - Clientes.xlsx         (Customers)
    - Vehiculos.xlsx        (Vehicles)
    - Ordenes de trabajo.xlsx (Work Orders)
    - productos.json        (Parts inventory)

Output:
    - app/src/main/assets/seed/seed_data.sql
"""

import os
import json
import hashlib
import base64
import secrets
from datetime import datetime
from collections import OrderedDict

import openpyxl

# ─── Configuration ────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "app", "src", "main", "assets", "seed", "seed_data.sql")

DATA_DIR = os.path.join(SCRIPT_DIR, "data")
CATALOGOS_FILE = os.path.join(DATA_DIR, "Catalogos.xlsx")
MARCAS_FILE = os.path.join(DATA_DIR, "Marcas y Modelos.xlsx")
CLIENTES_FILE = os.path.join(DATA_DIR, "Clientes.xlsx")
VEHICULOS_FILE = os.path.join(DATA_DIR, "Vehiculos.xlsx")
ORDENES_FILE = os.path.join(DATA_DIR, "Ordenes de trabajo.xlsx")
PRODUCTOS_FILE = os.path.join(DATA_DIR, "productos.json")

# Vehicle types for service duplication
SERVICE_VEHICLE_TYPES = ["SEDAN", "SUV", "CAMIONETA"]

DEFAULT_PASSWORD = "f4d3s2a1"

# ─── Password hashing (matches SecurityUtils.kt) ─────────────────────────────

def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    salt_b64 = base64.b64encode(salt).decode('utf-8')
    raw = f"{salt_b64}:{password}".encode('utf-8')
    h = hashlib.sha256(raw).hexdigest()
    return f"{salt_b64}:{h}"


def sql_escape(val):
    """Escape a value for SQL insertion."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s:
        return "NULL"
    s = s.replace("'", "''")
    s = ' '.join(s.split())
    return f"'{s}'"


def to_epoch_ms(dt_val) -> int:
    """Convert a datetime or string to epoch milliseconds."""
    if dt_val is None:
        return int(datetime.now().timestamp() * 1000)
    if isinstance(dt_val, datetime):
        return int(dt_val.timestamp() * 1000)
    if isinstance(dt_val, str):
        try:
            dt = datetime.strptime(dt_val.strip(), "%Y-%m-%d %H:%M:%S")
            return int(dt.timestamp() * 1000)
        except ValueError:
            pass
        try:
            dt = datetime.strptime(dt_val.strip(), "%Y-%m-%d")
            return int(dt.timestamp() * 1000)
        except ValueError:
            pass
    return int(datetime.now().timestamp() * 1000)


def read_single_column(ws, col_index=0):
    """Read a single column from a worksheet (skip header), return list of uppercase strings."""
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_index]
        if val is not None:
            s = str(val).strip().upper()
            if s:
                result.append(s)
    return result


# ─── Excel Readers ───────────────────────────────────────────────────────────

def read_catalogos(wb):
    """Read all catalog data from Catalogos.xlsx."""
    data = {}

    # Users (Nombre, Usuario, Rol, Contraseña)
    ws = wb['Usuarios']
    data['users'] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            name = str(row[0]).strip().upper()
            username = str(row[1]).strip() if row[1] else ""
            role = str(row[2]).strip().upper() if row[2] else "MECANICO"
            data['users'].append((name, username, role))

    # Vehicle Types
    ws = wb['Tipos de Vehiculo']
    data['vehicle_types'] = read_single_column(ws)

    # Colors
    ws = wb['Colores']
    data['colors'] = read_single_column(ws)

    # Accessories
    ws = wb['Accesorios']
    data['accessories'] = read_single_column(ws)

    # Part Brands
    ws = wb['Marcas de Repuestos']
    data['part_brands'] = read_single_column(ws)

    # Services (Categoria, Servicio, Precio Default)
    ws = wb['Servicios']
    data['services'] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            category = str(row[0]).strip().upper()
            service = str(row[1]).strip().upper()
            price = float(row[2]) if row[2] else 10.0
            if category not in data['services']:
                data['services'][category] = []
            data['services'][category].append((service, price))

    # Oil Types (Aceites)
    if 'Aceites' in wb.sheetnames:
        ws = wb['Aceites']
        data['oil_types'] = read_single_column(ws)
    else:
        data['oil_types'] = []

    # Motivos (complaints)
    if 'Motivos' in wb.sheetnames:
        ws = wb['Motivos']
        data['complaints'] = read_single_column(ws)
    else:
        data['complaints'] = []

    # Diagnosticos (Motivo, Diagnostico)
    if 'Diagnosticos' in wb.sheetnames:
        ws = wb['Diagnosticos']
        data['diagnoses'] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                complaint = str(row[0]).strip().upper()
                diagnosis = str(row[1]).strip().upper()
                data['diagnoses'].append((complaint, diagnosis))
    else:
        data['diagnoses'] = []

    return data


def read_brands_models(wb):
    """Read brands and models from Marcas y Modelos.xlsx, return OrderedDict."""
    ws = wb.active
    brands = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = str(row[0]).strip() if row[0] else None
        model = str(row[1]).strip() if row[1] and str(row[1]).strip() else None
        if not brand:
            continue
        if brand not in brands:
            brands[brand] = []
        if model:
            brands[brand].append(model)
    return brands


def read_parts(filepath):
    """Read parts from productos.json."""
    if not os.path.exists(filepath):
        print(f"  Warning: {filepath} not found, skipping parts")
        return []
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


# ─── Main Generator ──────────────────────────────────────────────────────────

def generate():
    lines = []
    lines.append("-- Serviaux Seed Data (auto-generated)")
    lines.append(f"-- Generated: {datetime.now().isoformat()}")
    lines.append("")

    now_ms = int(datetime.now().timestamp() * 1000)

    # ── Load Catalogos.xlsx ────────────────────────────────────────────────
    print("Reading Catalogos.xlsx...")
    wb_cat = openpyxl.load_workbook(CATALOGOS_FILE)
    catalogs = read_catalogos(wb_cat)

    # ── Load Marcas y Modelos.xlsx ─────────────────────────────────────────
    print("Reading Marcas y Modelos.xlsx...")
    wb_brands = openpyxl.load_workbook(MARCAS_FILE)
    brands_with_models = read_brands_models(wb_brands)

    # ── 1. Users ──────────────────────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- USERS")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for name, username, role in catalogs['users']:
        pw_hash = hash_password(DEFAULT_PASSWORD)
        lines.append(
            f"INSERT INTO users (name, username, role, passwordHash, commissionType, commissionValue, active, createdAt, updatedAt) "
            f"VALUES ({sql_escape(name)}, {sql_escape(username)}, '{role}', {sql_escape(pw_hash)}, 'NINGUNA', 0.0, 1, {now_ms}, {now_ms});"
        )
    lines.append("")

    # ── 2. Catalog: Vehicle Types ─────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: VEHICLE TYPES")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for vtype in catalogs['vehicle_types']:
        lines.append(f"INSERT INTO catalog_vehicle_types (name) VALUES ({sql_escape(vtype)});")
    lines.append("")

    # ── 3. Catalog: Colors ────────────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: COLORS")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for color in catalogs['colors']:
        lines.append(f"INSERT INTO catalog_colors (name) VALUES ({sql_escape(color)});")
    lines.append("")

    # ── 4. Catalog: Brands & Models ───────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: BRANDS & MODELS")
    lines.append("-- ══════════════════════════════════════════════════════════")
    brand_id = 0
    for brand_name, models in brands_with_models.items():
        brand_id += 1
        lines.append(f"INSERT INTO catalog_brands (id, name) VALUES ({brand_id}, {sql_escape(brand_name.upper())});")
        for model_name in models:
            lines.append(f"INSERT INTO catalog_models (brandId, name) VALUES ({brand_id}, {sql_escape(model_name.upper())});")
    lines.append("")

    # ── 5. Catalog: Part Brands ───────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: PART BRANDS")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for pb in catalogs['part_brands']:
        lines.append(f"INSERT INTO catalog_part_brands (name) VALUES ({sql_escape(pb)});")
    lines.append("")

    # ── 6. Catalog: Accessories ───────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: ACCESSORIES")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for acc in catalogs['accessories']:
        lines.append(f"INSERT INTO catalog_accessories (name) VALUES ({sql_escape(acc)});")
    lines.append("")

    # ── 7a. Catalog: Oil Types ──────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: OIL TYPES")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for oil in catalogs['oil_types']:
        lines.append(f"INSERT INTO catalog_oil_types (name) VALUES ({sql_escape(oil)});")
    lines.append("")

    # ── 7. Catalog: Services (duplicated per vehicle type) ──────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: SERVICES")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for category, services in catalogs['services'].items():
        for svc_name, price in services:
            for vtype in SERVICE_VEHICLE_TYPES:
                lines.append(
                    f"INSERT INTO catalog_services (category, name, defaultPrice, vehicleType) "
                    f"VALUES ({sql_escape(category)}, {sql_escape(svc_name)}, {price}, {sql_escape(vtype)});"
                )
    lines.append("")

    # ── 8. Catalog: Complaints (Motivos) ──────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: COMPLAINTS (MOTIVOS)")
    lines.append("-- ══════════════════════════════════════════════════════════")
    complaint_name_to_id = {}
    complaint_id = 0
    for complaint in catalogs['complaints']:
        complaint_id += 1
        complaint_name_to_id[complaint] = complaint_id
        lines.append(f"INSERT INTO catalog_complaints (id, name) VALUES ({complaint_id}, {sql_escape(complaint)});")
    lines.append("")

    # ── 9. Catalog: Diagnoses (Diagnósticos) ──────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CATALOG: DIAGNOSES (DIAGNOSTICOS)")
    lines.append("-- ══════════════════════════════════════════════════════════")
    for complaint_name, diagnosis in catalogs['diagnoses']:
        c_id = complaint_name_to_id.get(complaint_name, 0)
        if c_id > 0:
            lines.append(
                f"INSERT INTO catalog_diagnoses (complaintId, name) VALUES ({c_id}, {sql_escape(diagnosis)});"
            )
        else:
            print(f"  Warning: complaint '{complaint_name}' not found for diagnosis '{diagnosis}'")
    lines.append("")

    # ── 10. Parts from productos.json ─────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- PARTS (from productos.json)")
    lines.append("-- ══════════════════════════════════════════════════════════")
    print("Reading productos.json...")
    parts = read_parts(PRODUCTOS_FILE)
    for p in parts:
        nombre = str(p.get('nombre', '')).strip().upper()
        if not nombre:
            continue
        # Build description from all available fields
        desc_parts = []
        for field in ['descripcion', 'nombre_alterno', 'nombre_original', 'descripcion_etiqueta']:
            val = p.get(field)
            if val and str(val).strip() and str(val).strip().upper() not in ('SIN DEFINIR', ''):
                desc_parts.append(str(val).strip().upper())
        description = ' | '.join(desc_parts) if desc_parts else nombre

        codigo = str(p.get('codigo', '')).strip()
        precio = float(p.get('precio_prom', 0))
        stock = int(p.get('stock', 0))
        if stock < 0:
            stock = 0
        lines.append(
            f"INSERT INTO parts (name, description, code, brand, unitCost, salePrice, currentStock, active, createdAt, updatedAt) "
            f"VALUES ({sql_escape(nombre)}, {sql_escape(description)}, {sql_escape(codigo)}, 'GENERICO', {precio}, {precio}, {stock}, 1, {now_ms}, {now_ms});"
        )
    lines.append("")

    # ── 11. Customers from Excel ───────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- CUSTOMERS (from Clientes.xlsx)")
    lines.append("-- ══════════════════════════════════════════════════════════")

    # First insert "Consumidor Final" as customer ID 1
    lines.append(
        f"INSERT INTO customers (id, fullName, docType, idNumber, phone, email, address, notes, createdAt, updatedAt) "
        f"VALUES (1, 'CONSUMIDOR FINAL', 'CEDULA', '9999999999', '', NULL, NULL, 'CLIENTE GENERICO PARA VEHICULOS SIN PROPIETARIO ASIGNADO', {now_ms}, {now_ms});"
    )

    print("Reading Clientes.xlsx...")
    wb_clients = openpyxl.load_workbook(CLIENTES_FILE)
    ws_clients = wb_clients.active

    client_name_to_id = {}
    client_id_number_to_db_id = {}
    customer_id = 1  # Start after "Consumidor Final"

    for row in ws_clients.iter_rows(min_row=2, max_row=ws_clients.max_row, values_only=True):
        full_name = str(row[0]).strip().upper() if row[0] else None
        if not full_name:
            continue

        customer_id += 1
        id_number = str(row[1]).strip() if row[1] else None
        phone = str(row[2]).strip() if row[2] else ""
        created_at = to_epoch_ms(row[3])

        if phone:
            phone = phone.replace(".0", "")

        lines.append(
            f"INSERT INTO customers (id, fullName, docType, idNumber, phone, email, address, notes, createdAt, updatedAt) "
            f"VALUES ({customer_id}, {sql_escape(full_name)}, 'CEDULA', {sql_escape(id_number)}, {sql_escape(phone)}, NULL, NULL, NULL, {created_at}, {created_at});"
        )

        normalized = full_name.lower().strip()
        client_name_to_id[normalized] = customer_id
        if id_number:
            client_id_number_to_db_id[id_number] = customer_id

    lines.append("")

    # ── 12. Read Work Orders to build plate -> customer mapping ────────────
    print("Reading Ordenes de trabajo.xlsx...")
    wb_orders = openpyxl.load_workbook(ORDENES_FILE)
    ws_orders = wb_orders.active

    plate_to_customer_id = {}
    order_rows = []

    for row in ws_orders.iter_rows(min_row=2, max_row=ws_orders.max_row, values_only=True):
        order_id_raw = str(row[0]).strip() if row[0] else None
        if not order_id_raw:
            continue

        nombre = str(row[2]).strip() if row[2] else ""
        apellidos = str(row[3]).strip() if row[3] else ""
        plate = str(row[5]).strip() if row[5] else ""

        full_name_order = f"{nombre} {apellidos}".strip()
        normalized_order = full_name_order.lower().strip()

        matched_customer_id = None
        if normalized_order in client_name_to_id:
            matched_customer_id = client_name_to_id[normalized_order]
        else:
            for client_name, cid in client_name_to_id.items():
                if normalized_order in client_name or client_name in normalized_order:
                    matched_customer_id = cid
                    break

        if matched_customer_id and plate:
            plate_to_customer_id[plate.upper()] = matched_customer_id

        order_rows.append((row, matched_customer_id))

    print(f"  Mapped {len(plate_to_customer_id)} plates to customers")

    # ── 13. Vehicles from Excel ───────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- VEHICLES (from Vehiculos.xlsx)")
    lines.append("-- ══════════════════════════════════════════════════════════")

    print("Reading Vehiculos.xlsx...")
    wb_vehicles = openpyxl.load_workbook(VEHICULOS_FILE)
    ws_vehicles = wb_vehicles.active

    plate_to_vehicle_id = {}
    vehicle_id = 0

    for row in ws_vehicles.iter_rows(min_row=2, max_row=ws_vehicles.max_row, values_only=True):
        plate = str(row[6]).strip() if row[6] else None
        if not plate:
            continue

        vehicle_id += 1
        plate_upper = plate.upper()

        tipo = str(row[0]).strip().upper() if row[0] else None
        marca = str(row[1]).strip().upper() if row[1] else ""
        modelo = str(row[2]).strip().upper() if row[2] else ""
        anio = row[3]
        if anio:
            try:
                anio = int(float(str(anio)))
            except (ValueError, TypeError):
                anio = None
        version = str(row[4]).strip().upper() if row[4] else None
        color = str(row[5]).strip().upper() if row[5] else None
        transmision = str(row[7]).strip().upper() if row[7] else "MANUAL"
        motor = str(row[8]).strip().upper() if row[8] else None
        traccion = str(row[9]).strip().upper() if row[9] else "4X2"
        kms = row[10]
        if kms:
            try:
                kms = int(float(str(kms)))
            except (ValueError, TypeError):
                kms = None
        vin = str(row[11]).strip().upper() if row[11] else None
        num_motor = str(row[12]).strip().upper() if row[12] else None
        created_at = to_epoch_ms(row[13])

        cust_id = plate_to_customer_id.get(plate_upper, 1)

        year_sql = str(anio) if anio else "NULL"
        kms_sql = str(kms) if kms else "NULL"

        lines.append(
            f"INSERT INTO vehicles (id, customerId, plate, brand, model, version, year, vin, color, vehicleType, "
            f"currentMileage, engineDisplacement, engineNumber, drivetrain, transmission, notes, photoPaths, createdAt, updatedAt) "
            f"VALUES ({vehicle_id}, {cust_id}, {sql_escape(plate)}, {sql_escape(marca)}, {sql_escape(modelo)}, "
            f"{sql_escape(version)}, {year_sql}, {sql_escape(vin)}, {sql_escape(color)}, {sql_escape(tipo)}, "
            f"{kms_sql}, {sql_escape(motor)}, {sql_escape(num_motor)}, {sql_escape(traccion)}, {sql_escape(transmision)}, "
            f"NULL, NULL, {created_at}, {created_at});"
        )

        plate_to_vehicle_id[plate_upper] = vehicle_id

    lines.append("")

    # ── 14. Work Orders from Excel ────────────────────────────────────────
    lines.append("-- ══════════════════════════════════════════════════════════")
    lines.append("-- WORK ORDERS (from Ordenes de trabajo.xlsx)")
    lines.append("-- ══════════════════════════════════════════════════════════")

    admin_user_id = 1

    for row_data, matched_customer_id in order_rows:
        row = row_data
        order_id_raw = str(row[0]).strip() if row[0] else None
        if not order_id_raw:
            continue

        order_num = order_id_raw.replace("#", "").strip()
        try:
            order_id_int = int(order_num)
        except ValueError:
            continue

        entry_date = to_epoch_ms(row[1])
        falla = str(row[4]).strip().upper() if row[4] else "SIN DESCRIPCION"
        plate = str(row[5]).strip() if row[5] else ""
        plate_upper = plate.upper()

        kms = row[9]
        if kms:
            try:
                kms = int(float(str(kms)))
            except (ValueError, TypeError):
                kms = None

        total = row[13]
        try:
            total_val = float(total) if total else 0.0
        except (ValueError, TypeError):
            total_val = 0.0

        estado = str(row[14]).strip() if row[14] else "Entregado"

        estado_map = {
            "Entregado": "ENTREGADO",
            "Pendiente": "RECIBIDO",
            "En proceso": "EN_PROCESO",
            "En diagnóstico": "EN_DIAGNOSTICO",
            "En espera": "EN_ESPERA_REPUESTO",
            "Listo": "LISTO",
            "Cancelado": "CANCELADO",
        }
        status = estado_map.get(estado, "ENTREGADO")

        vehicle_id_ref = plate_to_vehicle_id.get(plate_upper)
        if not vehicle_id_ref:
            continue

        customer_id_ref = matched_customer_id or plate_to_customer_id.get(plate_upper, 1)

        kms_sql = str(kms) if kms else "NULL"

        lines.append(
            f"INSERT INTO work_orders (id, vehicleId, customerId, entryDate, status, priority, "
            f"orderType, arrivalCondition, "
            f"customerComplaint, initialDiagnosis, assignedMechanicId, entryMileage, fuelLevel, "
            f"checklistNotes, totalLabor, totalParts, total, photoPaths, createdBy, updatedBy, createdAt, updatedAt) "
            f"VALUES ({order_id_int}, {vehicle_id_ref}, {customer_id_ref}, {entry_date}, '{status}', 'MEDIA', "
            f"'SERVICIO_NUEVO', 'RODANDO', "
            f"{sql_escape(falla)}, NULL, NULL, {kms_sql}, NULL, "
            f"NULL, {total_val}, 0.0, {total_val}, NULL, {admin_user_id}, {admin_user_id}, {entry_date}, {entry_date});"
        )

    lines.append("")
    lines.append("-- End of seed data")

    # ── Write output ──────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"\nSeed data generated successfully!")
    print(f"Output: {OUTPUT_PATH}")

    # Print stats
    stats = {
        "Users": "INSERT INTO users",
        "Customers": "INSERT INTO customers",
        "Vehicles": "INSERT INTO vehicles",
        "Work Orders": "INSERT INTO work_orders",
        "Brands": "INSERT INTO catalog_brands",
        "Models": "INSERT INTO catalog_models",
        "Colors": "INSERT INTO catalog_colors",
        "Vehicle Types": "INSERT INTO catalog_vehicle_types",
        "Part Brands": "INSERT INTO catalog_part_brands",
        "Accessories": "INSERT INTO catalog_accessories",
        "Oil Types": "INSERT INTO catalog_oil_types",
        "Services": "INSERT INTO catalog_services",
        "Complaints": "INSERT INTO catalog_complaints",
        "Diagnoses": "INSERT INTO catalog_diagnoses",
        "Parts": "INSERT INTO parts",
    }

    print(f"\nStats:")
    for label, pattern in stats.items():
        count = sum(1 for l in lines if pattern in l)
        print(f"  {label:20s} {count}")


if __name__ == "__main__":
    generate()
